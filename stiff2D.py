# 変位法による平面骨組の応力解析
# 最初にクラス Stiff2D の定義があります
# これを呼び出すメインルーチンは本ファイルの末尾にあります

import math
import openpyxl
from openpyxl.utils import get_column_letter # 列幅の指定 2020/05/27

class Stiff2D():

    def __init__(self, inp_path='data.xlsx', out_path='result.xlsx'):
        #self.inp_path = inp_path # 入力ファイルのパス
        #self.out_path = out_path # 出力ファイルのパス
        self.node = []
        self.memb = []
        self.node_load = []
        self.memb_load = []
        self.node_index = [] # node_index[i] 節点番号iが定義されている node[]のindex
        self.memb_index = [] # memb_index[i] 部材番号iが定義されている memb[]のindex
        self.error = "" # エラーメッセージ

        
    def add_node(self, id = 0, x = 0, y = 0, xfix = 0, yfix = 0, rfix = 0):
        # 節点情報を一つずつ追加
        # id: 節点番号, x: X座標(cm), y: Y座標(cm), xfix:1の時、X方向固定
        #     yfix:1の時、Y方向固定, rfix:1の時、回転方向固定
        # node[i][j]
        # j = 0: id, 1: x, 2: y, 3:xfix, 4: yfix, 5: rfix, 
        #     計算結果 6: X方向変位(mm), 7: Y方向変位(mm), 8: 回転変位(rad)
        data = [0 for i in range(9)]
        data[0] = id
        data[1] = 100 * x # m -> cm
        data[2] = 100 * y
        data[3] = xfix
        data[4] = yfix
        data[5] = rfix
        self.node.append(data)
        
    def set_all_node(self, list2):
        # 節点情報を list2[][] から一括読み込み
        # 戻り値 0: 失敗, 1: 成功
        if len(list2) > 1: # list2[0]はヘッダー
            try: 
                for i in range(1, len(list2)):
                    data = [0 for i in range(9)]
                    for j in range(6):
                        if list2[i][j] is not None:
                            data[j] = list2[i][j]
                    data[1] *= 100 # m -> cm
                    data[2] *= 100 # m -> cm
                    self.node.append(data)
            except:
                return 0
        return 1
                
    def add_memb(self, id = 0, inode = 0, jnode = 0, a = 0, iz = 0, sa = 0,
                 e = 0, g = 0, ipin = 0, jpin = 0, irigid = 0, jrigid = 0):
        # 部材情報の追加 断面性能の単位 cm
        # id: 部材番号, inode: 始端の節点番号, jnode: 終端の節点番号, a:A, iz:I, 
        # sa:As, e:E, g:G, ipin:1の時、始端ピン, jpin:1の時、終端ピン,
        # irigid: 始端の剛域長, jrigid: 終端の剛域長
        # memb[i][j]
        # j = 0: id, 1: inode, 2: jnode, 3: a, 4: iz, 5: sa, 6: E, 7: G, 
        #     8: ipin, 9: jpin, 10: irigid, 11: jrigid,
        #     計算結果 12: Mi, 13: Mj, 14: Mc, 15: Qi, 16: Qj, 17: Ni, 18: Nj 
        data = [0 for i in range(19)]
        data[0] = id
        data[1] = inode
        data[2] = jnode
        data[3] = a
        data[4] = iz
        data[5] = sa
        data[6] = 0.1 * e # N/mm2 -> kN/cm2
        data[7] = 0.1 * g # N/mm2 -> kN/cm2
        data[8] = ipin
        data[9] = jpin
        data[10] = 100 * irigid # m -> cm
        data[11] = 100 * jrigid # m -> cm
        self.memb.append(data)
        
        
    def set_all_memb(self, list2):
        # 部材情報を list2[][] から一括読み込み
        # 戻り値 0: 失敗, 1: 成功
        if len(list2) > 1:  # list2[0]はヘッダー
            try: 
                for i in range(1, len(list2)):
                    data = [0 for i in range(19)]
                    for j in range(12):
                        if list2[i][j] is not None:
                            data[j] = list2[i][j]
                    data[6] *= 0.1 # N/mm2 -> kN/cm2
                    data[7] *= 0.1 # N/mm2 -> kN/cm2
                    data[10] *= 100 # m -> cm
                    data[11] *= 100 # m -> cm
                    self.memb.append(data)
            except:
                return 0
        return 1
        
    def add_node_load(self, num = 0, px = 0, py = 0, pr = 0):
        # 節点荷重情報の追加
        # num: 節点番号, px: X方向荷重(kN), py: Y方向荷重, pr:回転方向荷重(kN.m) 
        # node_load[i][j]
        # j = 0: num, 1: px, 2: py, 3: pr 
        data = [0 for i in range(4)]
        data[0] = num
        data[1] = px
        data[2] = py
        data[3] = -100 * pr # kN.m -> kN.cm 時計回りが正
        self.node_load.append(data)

    def set_all_node_load(self, list2):
        # 節点荷重情報を list2[][] から一括読み込み
        # 戻り値 0: 失敗, 1: 成功
        if len(list2) > 1: # list2[0]はヘッダー
            try: 
                for i in range(1, len(list2)):
                    data = [0 for i in range(4)]
                    for j in range(4):
                        if list2[i][j] is not None:
                            data[j] = list2[i][j]
                    data[3] *= -100 # kN.m -> kN.cm 時計回りが正
                    self.node_load.append(data)
            except:
                return 0
        return 1

    def add_memb_load(self, num = 0, ci = 0, cj = 0, m0 = 0, qi = 0, qj = 0, 
                      ni = 0, nj = 0):
        # 部材荷重情報の追加
        # num: 部材番号, ci: i端の固定端モーメント(kN.m), cj: j端の固定端モーメント(kN.m)
        # m0: 単純梁の中央モーメント(kN.m), qi: i端の固定端せん断力(kN), 
        # qj: j端の固定端せん断力(kN), ni: i端の固定端軸力(kN), nj: j端の固定端軸力(kN)
        # memb_load[i][j]
        # j = 0: num, 1: ci, 2: cj, 3: m0, 4: qi, 5: qj, 6: ni, 7: nj
        data = [0 for i in range(8)]
        data[0] = num
        data[1] = -100 * ci # kN.m -> kN.cm, 時計回りが正 -> 反時計回りが正
        data[2] = -100 * cj # kN.m -> kN.cm, 時計回りが正 -> 反時計回りが正
        data[3] = 100 * m0 # kN.m -> kN.cm
        data[4] = qi
        data[5] = qj
        data[6] = ni
        data[7] = nj
        self.memb_load.append(data)
        
    def set_all_memb_load(self, list2):
        # 部材荷重情報を list2[][] から一括読み込み
        # 戻り値 0: 失敗, 1: 成功
        if len(list2) > 1: # list2[0]はヘッダー
            try: 
                for i in range(1, len(list2)):
                    data = [0 for i in range(8)]
                    for j in range(8):
                        if list2[i][j] is not None:
                            data[j] = list2[i][j]
                    data[1] *= -100 # kN.m -> kN.cm 時計回りが正
                    data[2] *= -100 # kN.m -> kN.cm 時計回りが正
                    data[3] *= 100 # kN.m -> kN.cm
                    self.memb_load.append(data)
            except:
                return 0
        return 1

    # Start Calc def
    ########################################################################
    def calc_main(self):

        if not self.pre_check():
            return 0
        
        # 自由度(未知数)に番号をふる
        # free[i][j]: 節点node[i]の自由度番号 
        #         j = 0:X方向, 1:Y方向, 2:回転方向 
        free = []
        num_free = -1
        for i in range(len(self.node)):
            free1 = [-1, -1, -1]
            for j in range(3):
                if self.node[i][j + 3] == 0:
                    num_free += 1
                    free1[j] = num_free
            free.append(free1)
            
        max_free = num_free + 1 # 自由度(未知数)の数 = 全体剛性マトリクスのサイズ
        if max_free <= 0:
            self.error = "自由度が存在しない"
            return 0
        
        # バンドマトリクスの作成
        # counts[i]: バンドマトリクスのi列にある要素の数
        counts = [0 for i in range(max_free)]
        for m in range(len(self.memb)):
            for i in range(2): # i端/j端
                na = self.node_index[self.memb[m][i + 1]]
                for j in range(2):
                    nb = self.node_index[self.memb[m][j + 1]]
                    for k in range(3): # X/Y/R
                        for n in range(3): # X/Y/R
                            ndi = free[na][k]
                            ndj = free[nb][n]
                            if (ndi != -1) and (ndj != -1):
                                if (ndj - ndi + 1) > counts[ndj]:
                                    counts[ndj] = ndj - ndi + 1

        #  diag_num[i]: バンドマトリクスi列の対角要素の番号
        # 全体剛性マトリクスの要素[i,j]のバンドマトリクスの要素番号は diag_num[j]+j-i
        diag_num = [0 for i in range(max_free + 1)]
        num = 0
        for i in range(1, max_free + 1):
            num += counts[i - 1]
            diag_num[i] = num
        # 全体剛性マトリクスのバンドマトリクス band[]
        band = self.global_matrix(free, diag_num, max_free)
        # band[] を上三角行列に変換
        if not self.pre_gauss(band, diag_num, num_free):
            return 0
        # cmq_index[i] : memb[i] に荷重項の指定がある場合の memb_load[] のindex
        cmq_index = [-1 for i in range(len(self.memb))]
        # 外力ベクトル load[] の作成
        load = self.load_vector(free, num_free, cmq_index)
        # 変位ベクトル disp[] の計算
        disp = self.solve_gauss(band, diag_num, num_free, load)
        # 節点変位量を node[] に格納
        self.set_node_disp(free, disp)
        # 部材応力を memb[] に格納
        self.set_memb_stress(free, disp, cmq_index)
        return 1

    ########################################################################
    def global_matrix(self, free, diag_num, max_free):
        # 戻り値 全体剛性マトリクスのバンドマトリクス
        band = [0 for i in range(diag_num[max_free])] # 全体剛性マトリクスのバンドマトリクス
        for m in range(len(self.memb)):
            memb_stiff = self.memb_matrix(m, 1) # 部材剛性マトリクス [6][6]
            # 部材剛性マトリクス memb_stiff[][] を band[] に組み込む
            for i in range(2): # i端/j端
                na = self.node_index[self.memb[m][i + 1]]
                for j in range(2):
                    nb = self.node_index[self.memb[m][j + 1]]
                    for k in range(3): # X/Y/R
                        for n in range(3): # X/Y/R
                            ndi = free[na][k]
                            ndj = free[nb][n]
                            if (ndj >= ndi) and (ndi != -1) and (ndj != -1):
                                band[diag_num[ndj] + ndj - ndi] += \
                                        memb_stiff[3 * i + k][3 * j + n];
        return band        
            
    def memb_matrix(self, m, code):
        # 戻り値 memb[m]の部材マトリクス[6][6]
        # code = 0: 部材座標系, 1: 全体座標系
        # memb[][]のIndex
        (A, IZ, AS, E, G, PinI, PinJ, RgdI, RgdJ) = (3, 4, 5, 6, 7, 8, 9, 10, 11)
        free_num = 6; # 自由度数
        stiff =  [[0 for i in range(free_num)] for j in range(free_num)]
        (al, cs, sn) = self.memb_length(m)
        la = self.memb[m][RgdI]
        lb = self.memb[m][RgdJ]
        ld = al - la - lb # 可撓長
        ei = self.memb[m][E] * self.memb[m][IZ] # E*I
        rza = 1 # i端の接合状態による係数
        rzb = 1 # j端の接合状態による係数
        if self.memb[m][PinI] > 0:
            rza = 0
        if self.memb[m][PinJ] > 0:
            rzb = 0
        rys = 1 # せん断剛性の係数
        if (self.memb[m][AS] > 0.01) and (self.memb[m][G] > 0.01):
            ks = self.memb[m][G] * self.memb[m][AS] / ld
            rys = ks / (ks + 12 * ei / (ld * ld * ld))
        rz = rzb + 2 * rys + rzb * rys + rza * (1 - 4 * rzb * (rys - 1) + rys)
        kz11 = (12 * ei / (ld * ld * ld)) * rys * (rza + rzb + 4 * rza * rzb) / rz
        kz21 = (12 * ei / (ld * ld)) * rza * rys * (1 + 2 * rzb) / rz
        kz22 = (6 * ei / ld) * rza * (rzb + 2 * rys + rzb * rys) / rz
        kz41 = (12 * ei / (ld * ld)) * rzb * rys * (1 + 2 * rza) / rz
        kz42 = (6 * ei / ld) * rza * rzb * (-1 + 3 * rys) / rz
        kz44 = (6 * ei / ld) * rzb * (rza + 2 * rys + rza * rys) / rz
        # 部材座標系の剛性マトリクス
        stiff[0][0] = self.memb[m][E] * self.memb[m][A] / al
        stiff[1][1] = kz11
        stiff[2][2] = kz22 + 2 * kz21 * la + kz11 * la * la
        stiff[3][3] = stiff[0][0]
        stiff[4][4] = stiff[1][1]
        stiff[5][5] = kz44 + 2 * kz41 * lb + kz11 * lb * lb
        stiff[2][1] = kz21 + kz11 * la
        stiff[3][0] = -stiff[0][0]
        stiff[4][1] = -stiff[1][1]
        stiff[4][2] = -stiff[2][1]
        stiff[5][1] = kz41 + kz11 * lb
        stiff[5][2] = kz42 + kz11 * la * lb + kz41 * la + kz21 * lb
        stiff[5][4] = -stiff[5][1]
        stiff[1][2] = stiff[2][1]
        stiff[0][3] = stiff[3][0]
        stiff[1][4] = stiff[4][1]
        stiff[2][4] = stiff[4][2]
        stiff[1][5] = stiff[5][1]
        stiff[2][5] = stiff[5][2]
        stiff[4][5] = stiff[5][4]
        if code == 1:
            # 全体座標系に変換
            trm = self.trans_matrix(cs, sn) # 座標変換マトリクス [6][6]
            wrk = [[0 for i in range(free_num)] for j in range(free_num)]
            # 座標変換 [stiif] = [trm]T * [stiff] * [trm]
            for i in range(free_num):
                for j in range(free_num):
                    for k in range(free_num):
                        wrk[i][j] += stiff[i][k] * trm[k][j]
            for i in range(free_num):
                for j in range(free_num):
                    for k in range(free_num):
                        stiff[i][j] = 0
            for i in range(free_num):
                for j in range(free_num):
                    for k in range(free_num):
                        stiff[i][j] += trm[k][i] * wrk[k][j]
        return stiff    
            
    def trans_matrix(self, cs, sn):
        # 戻り値 部材の座標変換マトリクス [6][6]: 
        # cs: cosθ, sn: sinθ (θは部材の傾斜)
        trans = [[0 for i in range(6)] for j in range(6)]
        trans[0][0] = cs
        trans[0][1] = sn
        trans[1][0] = -sn
        trans[1][1] = cs
        trans[2][2] = 1.0
        trans[3][3] = cs
        trans[3][4] = sn
        trans[4][3] = -sn
        trans[4][4] = cs
        trans[5][5] = 1.0
        return trans
    
    def pre_gauss(self, a, diag_num, num_max):
        # バンドマトリクス a[] を上三角マトリクスに変換
        # diag_num[i]: a[] のi列の対角要素の番号
        # num_max: a[] の列数 - 1
        # 戻り値 0: 異常終了, 1: 正常終了
        mm = [0 for i in range(num_max + 1)]
        for i in range(num_max + 1):
            if a[diag_num[i]] < 0.01:
                self.error = "剛性マトリクスの対角要素が0になる"
                return 0
        # 上三角マトリクスの作成
        for j in range(num_max + 1):
            mm[j] = j - diag_num[j + 1] + diag_num[j] + 1
            nst = mm[j] + 1
            if j > nst:
                for i in range(nst, j):
                    mk1 = mm[i]
                    if mm[j] > mm[i]:
                        mk1 = mm[j]
                    mk2 = i - 1
                    if num_max + 1 < i:
                        mk2 = num_max
                    nij = diag_num[j] + j - i
                    for k in range(mk1, mk2 + 1):
                        a[nij] -= a[diag_num[i] + i - k] * a[diag_num[j] + j - k]
            njj = diag_num[j]
            mk2 = j - 1
            if num_max + 1 < j:
                mk2 = num_max
            if mk2 >= mm[j]:
                for i in range(mm[j], mk2 + 1):
                    nij = diag_num[j] + j - i
                    a[njj] -= a[nij] * a[nij] / a[diag_num[i]]
                    a[nij] /= a[diag_num[i]]
        return 1
    
    def solve_gauss(self, a, diag_num, num_max, r):
        # ガウス消去法による解法
        # a[]: 上三角マトリクス, diag_num[i]: a[] のi列の対角要素の番号
        # num_max: a[] の列数 - 1, r[]: 係数ベクトル
        # 戻り値 result[]: 解のベクトル
        result = []
        for i in range(len(r)):
            result.append(r[i])        
        mm = [0 for i in range(num_max + 1)]
        mm[0] = 1
        for i in range(1, num_max + 1):
            mm[i] = i - diag_num[i + 1] + diag_num[i] + 1
        for i in range(1, num_max + 1):
            for k in range(mm[i], i):
                n = diag_num[i] + i - k
                result[i] -= a[n] * result[k]
        # 2020/03/23
        # for i in range(1, num_max + 1):
        for i in range(num_max + 1):
            result[i] /= a[diag_num[i]]                 
        for i in reversed(range(0, num_max + 1)):
            for k in range(mm[i], i):
                n = diag_num[i] + i - k
                result[k] -= a[n] * result[i]
        return result                
 
    def load_vector(self, free, num_free, cmq_index):
        # 外力ベクトルの作成
        # free[i][j]: 節点node[i]の自由度番号, num_free: 自由度の数
        # cmq_index[i]: memb[i] に対応した memb_load[] の index
        # 戻り値 外力ベクトル vect[]
        vect = [0 for i in range(num_free + 1)]
        # 節点荷重による外力
        for i in range(len(self.node_load)):
            free_num = -1
            for j in range(3): # X/Y/R
                if abs(self.node_load[i][j + 1]) > 0.01:
                    num = self.node_index[self.node_load[i][0]]
                    free_num = free[num][j]
                    if free_num != -1:
                        vect[free_num] += self.node_load[i][j + 1]
                        
        # 部材荷重による外力
        for i in range(len(self.memb_load)):
            # 2020/05/12
            if self.memb_load[i][0] == 0:
                continue # 重複指定がある
            m = self.memb_index[self.memb_load[i][0]]
            cmq_index[m] = i
            cmq = self.cmq_vector(i, m)
            (al, cs, sn) = self.memb_length(m)
            # cs: cosθ, sn: sinθ
            # fem[]: 全体座標系に変換された cmq[] の成分
            fem = [0 for i in range(6)]
            fem[0] = -cmq[1] * sn + cmq[0] * cs # Xi
            fem[1] =  cmq[1] * cs + cmq[0] * sn # Yi
            fem[2] =  cmq[2] # Ri
            fem[3] = -cmq[4] * sn + cmq[3] * cs # Xj
            fem[4] =  cmq[4] * cs + cmq[3] * sn # Yj
            fem[5] =  cmq[5] # Rj
            ni = self.node_index[self.memb[m][1]] # i端節点の node[] の index
            nj = self.node_index[self.memb[m][2]] # j端節点の node[] の index
            # num[]: 両端の自由度 (Xi, Yi, Ri, Xj, Yj, Rj) の自由度番号
            num = [0 for i in range(6)]
            for j in range(3):
                num[j] = free[ni][j]
                num[3 + j] = free[nj][j]
            # 外力ベクトルに加算
            for j in range(6):
                if num[j] > -1:
                    vect[num[j]] -= fem[j];
            
        return vect
    
    def cmq_vector(self, n, m):
        # memb[m] memb_load[n] の荷重項を変形成分順に並べ替える
        # 戻り値 vect[i] i = 0:Ni, 1:Qi, 2:Mi, 3:Nj, 4:Qj, 5:Mj, 6:M0
        vect = [0 for i in range(7)]
        # memb_load[][] の index
        (Ci, Cj, M0, Qi, Qj, Ni, Nj) = (1, 2, 3, 4, 5, 6, 7)
        vect[0] = self.memb_load[n][Ni]
        vect[1] = self.memb_load[n][Qi]
        vect[2] = self.memb_load[n][Ci]
        vect[3] = self.memb_load[n][Nj]
        vect[4] = self.memb_load[n][Qj]
        vect[5] = self.memb_load[n][Cj]
        vect[6] = self.memb_load[n][M0]
        # 接合部がピンの場合
        if (self.memb[m][8] + self.memb[m][9] > 0):
            al = self.memb_length(m, 1)
            if (self.memb[m][8] == 1) and (self.memb[m][9] == 1):
                # 両端ピン
                vect[1] = vect[1] - (vect[2] + vect[5]) / al
                vect[4] = vect[4] + (vect[2] + vect[5]) / al
                vect[2] = 0
                vect[5] = 0
            elif self.memb[m][8] == 1:
                # i端ピン
                vect[5] = vect[5] - 0.5 * vect[2]
                vect[1] = vect[1] - 1.5 * vect[2] / al
                vect[4] = vect[4] + 1.5 * vect[2] / al
                vect[2] = 0
            elif self.memb[m][9] == 1:
                # j端ピン
                vect[2] = vect[2] - 0.5 * vect[5]
                vect[1] = vect[1] - 1.5 * vect[5] / al
                vect[4] = vect[4] + 1.5 * vect[5] / al
                vect[5] = 0

        return vect
    
    def set_node_disp(self, free, disp):
        # 節点変位量を node[] に格納
        # free[][]: 自由度番号の配列, disp[]: 全自由度の変位ベクトル
        (Dx, Dy, Dr) = (6, 7, 8)
        for i in range(len(self.node)):
            if free[i][0] > -1: # X
                self.node[i][Dx] = 10 * disp[free[i][0]] # cm -> mm
            if free[i][1] > -1:
                self.node[i][Dy] = 10 * disp[free[i][1]] # cm -> mm
            if free[i][2] > -1:
                self.node[i][Dr] = -disp[free[i][2]] # 時計回りが正

    def set_memb_stress(self, free, disp, cmq_index):
        # 部材応力を memb[] に格納
        # free[][]: 自由度番号の配列, disp[]: 全自由度の変位ベクトル
        # cmq_index[i]: memb[i] に対応した memb_load[] の index
        free_num = 6 # 自由度の数
        # memb[][] の index 
        (Mi, Mj, Mc, Qi, Qj, Ni, Nj) = (12, 13, 14, 15, 16, 17, 18)
        for m in range(len(self.memb)):
            # num[]: 部材両端の節点の自由度番号 (xi, yi, ri, xj, yj, rj)
            num = [0 for i in range(free_num)]
            for i in range(3): # X/Y/R
                num[i] = free[self.node_index[self.memb[m][1]]][i] # i端側
                num[3 + i] = free[self.node_index[self.memb[m][2]]][i] # j端側
            # global_disp[]: 部材両端の全体座標系の変位 (xi, yi, ri, xj, yj, rj)
            global_disp = [0 for i in range(free_num)]
            for i in range(free_num):
                if num[i] > -1:
                    global_disp[i] = disp[num[i]]
            # local_disp[]: 部材両端の部材座標系の変位 (xi, yi, ri, xj, yj, rj)
            local_disp = [0 for i in range(free_num)]
            # global_disp -> local_disp
            (al, cs, sn) = self.memb_length(m) # cs: cosθ, sn: sinθ
            local_disp[0] = global_disp[0] * cs + global_disp[1] * sn
            local_disp[1] = global_disp[1] * cs - global_disp[0] * sn
            local_disp[2] = global_disp[2]
            local_disp[3] = global_disp[3] * cs + global_disp[4] * sn
            local_disp[4] = global_disp[4] * cs - global_disp[3] * sn
            local_disp[5] = global_disp[5]
            # local_stiff[6][6]: 部材座標系の剛性マトリクス
            local_stiff = self.memb_matrix(m, 0)
            # stress[]: 部材応力 (xi, yi, ri, xj, yj, rj) = local_disp * local_stiff
            stress = [0 for i in range(free_num)]
            for i in range(free_num):
                for j in range(free_num):
                    stress[i] += local_stiff[i][j] * local_disp[j]
            m0 = 0 # 単純梁の中央曲げ
            if (cmq_index[m] > -1):
                # 荷重項の値を部材応力に加算
                cmq = self.cmq_vector(cmq_index[m], m)
                for i in range(free_num):
                    stress[i] += cmq[i]
                m0 = cmq[6]
            # 時計回りの曲げを正に変換
            stress[2] *= -1.0
            stress[5] *= -1.0
            self.memb[m][Mi] = 0.01 * stress[2] # kN.cm -> kN.m 
            self.memb[m][Mj] = 0.01 * stress[5] # kN.cm -> kN.m 
            self.memb[m][Mc] = 0.01 * (m0 + 0.5 * ((stress[2] - stress[5]))) 
            self.memb[m][Qi] = stress[1] 
            self.memb[m][Qj] = stress[4] 
            self.memb[m][Ni] = stress[0] 
            self.memb[m][Nj] = stress[3]

    def pre_check(self):
        # データの検証
        # 戻り値 0: エラーあり, 1: エラーなし
        if len(self.node) == 0:
            self.error = "節点情報がない"
            return 0
        elif len(self.memb) == 0:
            self.error = "部材情報がない"
            return 0
        elif (len(self.node_load) == 0) and (len(self.memb_load) == 0): 
            self.error = "荷重情報がない"
            return 0

        num_node = 0 # 節点番号の最大値
        for i in range(len(self.node)):
            if self.node[i][0] < 1:
                self.error = "節点番号" + str(self.node[i][0]) + "に誤りがある"
                return 0
            elif self.node[i][0] > num_node:
                num_node = self.node[i][0]
        # node_index[i] 節点番号iが定義されている node[]のindex
        self.node_index = [-1 for i in range(num_node + 1)]
        for i in range(len(self.node)):
            self.node_index[self.node[i][0]] = i

        num_memb = 0 # 部材番号の最大値
        for i in range(len(self.memb)):
            err_memb = "部材番号" + str(self.memb[i][0])
            if self.memb[i][0] < 1:
                self.error = err_memb + "に誤りがある"
                return 0
            elif (self.memb[i][1] < 1) or (self.memb[i][1] > num_node) or \
                    (self.memb[i][2] < 1) or (self.memb[i][2] > num_node):
                self.error = err_memb + "の節点情報に誤りがある"
                return 0
            elif (self.memb[i][3] < 0.1) or (self.memb[i][4] < 0.1):
                self.error = err_memb + "の断面性能に誤りがある"
                return 0
            elif self.memb[i][6] < 0.1:
                self.error = err_memb + "の材料定数に誤りがある"
                return 0
            else:
                al = self.memb_length(i, 1)
                if al < 0.01:
                    self.error = err_memb + "の長さが確保できない"
                    return 0
                elif al < (self.memb[i][10] + self.memb[i][11]):
                    self.error = err_memb + "剛域長の値に誤りがある"
                    return 0
                else:
                    if self.memb[i][0] > num_memb:
                        num_memb = self.memb[i][0]
                
        # memb_index[i] 部材番号iが定義されている memb[]のindex
        self.memb_index = [-1 for i in range(num_memb + 1)]
        for i in range(len(self.memb)):
            self.memb_index[self.memb[i][0]] = i
        # 節点荷重情報
        for i in range(len(self.node_load)):
            if (self.node_load[i][0] < 1) or (self.node_load[i][0] > num_node):
                self.error = "節点荷重が作用する節点番号" + \
                            str(self.node_load[i][0]) + "に誤りがある"
                return 0
        # 部材荷重情報
        for i in range(len(self.memb_load)):
            if (self.memb_load[i][0] < 1) or (self.memb_load[i][0] > num_memb):
                self.error = "部材荷重が作用する部材番号" + \
                            str(self.memb_load[i][0]) + "に誤りがある"
                return 0
        # 2020/05/12
        # 部材荷重が重複して指定されている場合は値を加算
        for i in range(len(self.memb_load)):
            if self.memb_load[i][0] != 0: 
                for j in range(i + 1, len(self.memb_load)):
                    if self.memb_load[i][0] == self.memb_load[j][0]:
                        self.memb_load[j][0] = 0
                        for k in range(8):
                            self.memb_load[i][k] += self.memb_load[j][k]
        return 1
               
    def memb_length(self, m, result = 0):
        # memb[m]の部材長
        # result = 1: 部材長のみを返す
        # 戻り値 al: 部材長, cs: cosθ, sn: sinθ
        x1 = self.node[self.node_index[self.memb[m][1]]][1]
        y1 = self.node[self.node_index[self.memb[m][1]]][2]
        x2 = self.node[self.node_index[self.memb[m][2]]][1]
        y2 = self.node[self.node_index[self.memb[m][2]]][2]
        al = (x2 - x1) ** 2 + (y2 - y1) ** 2
        cs = 0
        sn = 0
        if al > 0.01:
            al = math.sqrt(al)
            cs = (x2 - x1) / al
            sn = (y2 - y1) / al
        if result:
            return al
        else:
            return (al, cs, sn)
        
    def read_data_xlsx(self,inputFile):
        # data.xlsx からデータを読み込む
        # 戻り値 0: 失敗, 1: 成功
        try:
            wb = openpyxl.load_workbook(inputFile)
            self.set_all_node(list(wb['NODE'].values))
            self.set_all_memb(list(wb['MEMB'].values))
            self.set_all_node_load(list(wb['NodeLoad'].values))
            self.set_all_memb_load(list(wb['MemberLoad'].values))
            return 1
        except Exception as err:
            print(err)
            return 0

    def write_model(self):
        ########################################################################
        # Make data for the diagram

        # data for diagram
        self.xp = [[]]
        self.yp = [[]]

        for i in range(len(self.memb)):

            #print(self.memb[i][0],self.memb[i][1],self.memb[i][2])
            xi = []
            yi = []

            for j in range(len(self.node)):
                #print(self.node[j][0],self.node[j][1],self.node[j][2])

                if(self.memb[i][1] == self.node[j][0]):
                    xi.append(float(self.node[j][1])/100)
                    yi.append(float(self.node[j][2])/100)

                if(self.memb[i][2] == self.node[j][0]):
                    xi.append(float(self.node[j][1])/100)
                    yi.append(float(self.node[j][2])/100)

            self.xp.append(xi)
            self.yp.append(yi)

        # Check data
        #print(self.xp,self.yp)

        # End
        ########################################################################

    def write_result_xlsx(self,outputFile):
        # result.xlsx に計算結果を書き込む
        # 戻り値 0: 失敗, 1: 成功
        def result_value(f, n):
            # 浮動小数 f を小数以下 n 桁の値に変換
            # 2020/05/12 math.float(切り捨て) -> round(四捨五入)
            v = round((10 ** n) * abs(f)) / (10 ** n)
            if f > 0:
                return v
            else:
                return -v
        try:
            wb = openpyxl.Workbook()
            # 節点変位
            ws1 = wb.active
            ws1.title = 'Displacement'
            header1 = ['Node Num.', 'XDir,Disp.(mm)', 'YDir,Disp.(mm)', 'Rotation.(rad)']
            for i in range(len(header1)):
                ws1.cell(row = 1, column = i + 1).value = header1[i]
            for i in range(len(self.node)):
                result = []
                result.append(self.node[i][0])
                result.append(result_value(self.node[i][6], 2))
                result.append(result_value(self.node[i][7], 2))
                result.append(result_value(self.node[i][8], 4))
                for j in range(len(result)):
                    ws1.cell(row = i + 2, column = j + 1).value = result[j]
            # 列幅の指定 2020/05/27
            ws1.column_dimensions[get_column_letter(1)].width = 10
            for x in range(1, len(header1)):
                ws1.column_dimensions[get_column_letter(x + 1)].width = 15
            # 部材応力
            ws2 =  wb.create_sheet('Member Stress')
            header2 = ['Node Num.', 'Mi (kN.m)', 'Mj (kN.m)', 'Mc (kN.m)',
                       'Qi (kN)', 'Qj (kN)', 'Ni (kN)', 'Nj (kN)']
            for i in range(len(header2)):
                ws2.cell(row = 1, column = i + 1).value = header2[i]
            for i in range(len(self.memb)):
                result = []
                result.append(self.memb[i][0])
                for j in range(12, 19):
                    result.append(result_value(self.memb[i][j], 2))
                for j in range(len(result)):
                    ws2.cell(row = i + 2, column = j + 1).value = result[j]
            # 列幅の指定 2020/05/27
            ws2.column_dimensions[get_column_letter(1)].width = 10
            for x in range(1, len(header2)):
                ws2.column_dimensions[get_column_letter(x + 1)].width = 12
                wb.save(outputFile)
            return 1        
        except Exception as err:
            print(err)
            return 0

